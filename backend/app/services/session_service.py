import asyncio
import csv
import io
import os
import re
import sqlite3
import tempfile
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session, noload, object_session

from app.core.cache import redis_client
from app.core.config import get_settings
from app.core.database import SessionLocal
from app.models.customer import Customer
from app.models.message import Message
from app.models.session import SessionGroup, SessionLog, SessionStatus, SessionTaskLog, TelegramSession
from app.services.proxy_service import proxy_service
from app.services.target_parser import parse_targets
from app.services.websocket_manager import session_ws_manager
from app.services.session_command_bus import session_command_bus


settings = get_settings()


class SessionService:
    _CONNECT_QUEUE_LOCK_SECONDS = 180
    _BIDIRECTIONAL_QUEUE_LOCK_SECONDS = 300
    _CONTACT_SCAN_QUEUE_LOCK_SECONDS = 1800

    def __init__(self) -> None:
        # Keep strong references until completion; otherwise fire-and-forget
        # asyncio tasks may be garbage-collected while a batch is running.
        self._connect_tasks: set[asyncio.Task[Any]] = set()
        self._connect_semaphore = asyncio.Semaphore(max(settings.session_worker_count, 12))
        self._bidirectional_tasks: set[asyncio.Task[Any]] = set()
        self._bidirectional_semaphore = asyncio.Semaphore(max(settings.session_worker_count, 12))
        self._contact_scan_tasks: set[asyncio.Task[Any]] = set()
        self._contact_scan_semaphore = asyncio.Semaphore(max(settings.session_worker_count, 12))

    _SPAM_BOT_NORMAL_MARKERS = (
        "good news",
        "no limits",
        "free as a bird",
    )
    _SPAM_BOT_BLOCKED_MARKERS = (
        "your account was blocked",
        "your account has been blocked",
        "your account is blocked",
        "account was banned",
        "account has been banned",
    )
    _SPAM_BOT_RESTRICTED_MARKERS = (
        "limited",
        "restriction",
        "spam",
        "cannot send",
        "can't send",
        "too many",
        "deactivated",
        "prevented",
        "complaints",
    )

    def list_sessions(
        self,
        db: Session,
        group_id: int | None = None,
        kf_id: int | None = None,
        status: str | None = None,
        health_status: str | None = None,
        keyword: str | None = None,
        owner_id: int | None = None,
        bidirectional_status: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[TelegramSession], int]:
        filters = []
        if owner_id is not None:
            filters.append(TelegramSession.owner_id == owner_id)
        if group_id is not None:
            if group_id == 0:
                filters.append(TelegramSession.group_id.is_(None))
            else:
                filters.append(TelegramSession.group_id == group_id)
        if kf_id is not None:
            if kf_id == 0:
                filters.append(TelegramSession.kf_id.is_(None))
            else:
                filters.append(TelegramSession.kf_id == kf_id)
        if status:
            filters.append(TelegramSession.status == status)
        if health_status:
            filters.append(TelegramSession.health_status == health_status)
        if bidirectional_status:
            filters.append(TelegramSession.bidirectional_status == bidirectional_status)
        if keyword:
            like = f"%{keyword}%"
            filters.append(
                or_(
                    TelegramSession.phone.like(like),
                    TelegramSession.username.like(like),
                    TelegramSession.session_name.like(like),
                )
            )
        total = int(db.scalar(select(func.count()).select_from(TelegramSession).where(*filters)) or 0)
        stmt = (
            select(TelegramSession)
            .options(noload(TelegramSession.logs))
            .where(*filters)
            .order_by(TelegramSession.created_at.asc(), TelegramSession.id.asc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
        return list(db.scalars(stmt).all()), total

    def count_sent_messages_batch(self, db: Session, session_ids: list[int]) -> dict[int, int]:
        if not session_ids:
            return {}
        rows = db.execute(
            select(SessionTaskLog.session_id, func.count(SessionTaskLog.id))
            .where(SessionTaskLog.session_id.in_(session_ids), SessionTaskLog.status == "success")
            .group_by(SessionTaskLog.session_id)
        ).all()
        return {int(session_id): int(count) for session_id, count in rows if session_id is not None}

    def create_session(self, db: Session, data: dict[str, Any], owner_id: int | None = None) -> TelegramSession:
        session = TelegramSession(
            owner_id=owner_id,
            username=data["username"],
            avatar=data.get("avatar"),
            phone=data["phone"],
            session_name=data.get("session_name") or self._make_session_name(data["phone"]),
            group_id=data.get("group_id"),
        )
        db.add(session)
        db.flush()
        self.log(db, session.id, "create", "Session created")
        db.commit()
        db.refresh(session)
        return session

    async def update_session(self, db: Session, session_id: int, data: dict[str, Any], owner_id: int | None = None) -> TelegramSession:
        session = db.get(TelegramSession, session_id)
        if not session or (owner_id is not None and session.owner_id != owner_id):
            raise ValueError("Session not found")

        action = data.pop("action", None)
        for key, value in data.items():
            if hasattr(session, key) and value is not None:
                setattr(session, key, value)

        if action == "connect":
            await self.connect_session(db, session)
        elif action == "disconnect":
            await self.disconnect_session(db, session)
        else:
            self.log(db, session.id, "update", "Session updated")
            db.commit()
            db.refresh(session)
            await self.publish_status(session, "updated")
        return session

    async def connect_session(
        self,
        db: Session,
        session: TelegramSession,
        *,
        already_queued: bool = False,
    ) -> TelegramSession:
        runtime_online = bool(await redis_client.exists(f"telegram:session:runtime:{session.id}"))
        if (session.status == SessionStatus.connecting and not already_queued) or runtime_online:
            self.log(db, session.id, "connect_skipped", f"Session is already {session.status.value}")
            db.commit()
            db.refresh(session)
            return session
        if await redis_client.exists(f"marketing:session_lock:{session.id}"):
            self.log(db, session.id, "connect_skipped", "Session is busy sending a task")
            db.commit()
            db.refresh(session)
            return session
        session.status = SessionStatus.connecting
        session.error_message = None
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "status_changed")

        try:
            await session_command_bus.execute(session.id, "connect", timeout=60)
            session.status = SessionStatus.connected
            session.last_login_at = datetime.utcnow()
            session.health_status = "healthy"
            session.error_message = None
            self.log(db, session.id, "connect", "Session connected")
        except Exception as exc:
            session.status = SessionStatus.error
            session.health_status = "unhealthy"
            session.error_message = str(exc)
            self.log(db, session.id, "connect_failed", str(exc))

        db.commit()
        db.refresh(session)
        await self.publish_status(session, "status_changed")
        return session

    async def connect_sessions(
        self,
        db: Session,
        session_ids: list[int],
        owner_id: int | None = None,
    ) -> dict[str, Any]:
        sessions = self.get_sessions_by_ids(db, session_ids, owner_id)
        summary = {
            "requested": len(dict.fromkeys(session_ids)),
            "found": len(sessions),
            "accepted": 0,
            "accepted_ids": [],
            "skipped": 0,
        }
        for session in sessions:
            db.refresh(session)
            runtime_online = bool(await redis_client.exists(f"telegram:session:runtime:{session.id}"))
            lock_key = f"telegram:session:connect-pending:{session.id}"
            lock_token = uuid.uuid4().hex
            acquired = await redis_client.set(
                lock_key,
                lock_token,
                nx=True,
                ex=self._CONNECT_QUEUE_LOCK_SECONDS,
            )
            if session.status == SessionStatus.connecting or runtime_online or not acquired:
                if acquired:
                    await self._release_connect_queue_lock(lock_key, lock_token)
                summary["skipped"] += 1
                self.log(db, session.id, "batch_connect_skipped", f"Session is already {session.status.value}")
                db.commit()
                continue

            if await redis_client.exists(f"marketing:session_lock:{session.id}"):
                await self._release_connect_queue_lock(lock_key, lock_token)
                summary["skipped"] += 1
                self.log(db, session.id, "batch_connect_skipped", "Session is busy sending a task")
                db.commit()
                continue

            session.status = SessionStatus.connecting
            session.error_message = None
            self.log(db, session.id, "batch_connect_queued", "Session connection queued")
            db.commit()
            db.refresh(session)
            await self.publish_status(session, "status_changed")

            task = asyncio.create_task(
                self._run_queued_connect(session.id, lock_key, lock_token),
                name=f"connect-session-{session.id}",
            )
            self._connect_tasks.add(task)
            task.add_done_callback(self._connect_tasks.discard)
            summary["accepted"] += 1
            summary["accepted_ids"].append(session.id)
        return summary

    async def _run_queued_connect(self, session_id: int, lock_key: str, lock_token: str) -> None:
        try:
            async with self._connect_semaphore:
                db = SessionLocal()
                try:
                    session = db.get(TelegramSession, session_id)
                    if session and session.status == SessionStatus.connecting:
                        await self.connect_session(db, session, already_queued=True)
                finally:
                    db.close()
        finally:
            await self._release_connect_queue_lock(lock_key, lock_token)

    async def _release_connect_queue_lock(self, key: str, token: str) -> None:
        await redis_client.eval(
            "if redis.call('get', KEYS[1]) == ARGV[1] then "
            "return redis.call('del', KEYS[1]) else return 0 end",
            1,
            key,
            token,
        )

    async def disconnect_session(self, db: Session, session: TelegramSession) -> TelegramSession:
        if await redis_client.exists(f"marketing:session_lock:{session.id}"):
            raise ValueError("Session is busy sending a task")
        session.status = SessionStatus.disconnected
        session.health_status = "unknown"
        self.log(db, session.id, "disconnect", "Session disconnected")
        db.commit()
        db.refresh(session)
        await session_command_bus.execute(session.id, "disconnect", timeout=20)
        await self.publish_status(session, "status_changed")
        return session

    async def disconnect_sessions(
        self,
        db: Session,
        session_ids: list[int],
        owner_id: int | None = None,
    ) -> int:
        sessions = self.get_sessions_by_ids(db, session_ids, owner_id)
        sessions = [
            session for session in sessions
            if not await redis_client.exists(f"marketing:session_lock:{session.id}")
        ]
        for session in sessions:
            session.status = SessionStatus.disconnected
            session.health_status = "unknown"
            self.log(db, session.id, "batch_disconnect", "Session disconnected by batch operation")
        db.commit()

        for session in sessions:
            try:
                await session_command_bus.execute(session.id, "disconnect", timeout=20)
            except Exception:
                pass
            db.refresh(session)
            await self.publish_status(session, "status_changed")
        return len(sessions)

    async def delete_session(self, db: Session, session_id: int, owner_id: int | None = None) -> None:
        session = db.get(TelegramSession, session_id)
        if not session or (owner_id is not None and session.owner_id != owner_id):
            raise ValueError("Session not found")
        if await redis_client.exists(f"marketing:session_lock:{session.id}"):
            raise ValueError("Session is busy sending a task")
        if await redis_client.exists(f"telegram:session:owner:{session.id}"):
            try:
                await session_command_bus.execute(session.id, "disconnect", timeout=20)
            except Exception as exc:
                raise ValueError(f"无法安全释放Session Worker: {exc}") from exc
            for _ in range(20):
                if not await redis_client.exists(f"telegram:session:owner:{session.id}"):
                    break
                await asyncio.sleep(0.5)
            if await redis_client.exists(f"telegram:session:owner:{session.id}"):
                raise ValueError("Session Worker尚未释放文件，请稍后重试")
        session_file = Path(settings.session_dir) / f"{session.session_name}.session"
        # Delete both the message rows and their conversation-list entries.
        # Existing databases may either lack cascades or set the customer's
        # assigned_session_id to NULL, which would leave stale conversations.
        db.execute(delete(Message).where(Message.session_id == session.id))
        db.execute(delete(Customer).where(Customer.assigned_session_id == session.id))
        db.delete(session)
        db.commit()
        if session_file.exists():
            session_file.unlink()
        await session_ws_manager.broadcast("*", {"event": "deleted", "id": session_id})

    def list_groups(self, db: Session, owner_id: int | None = None) -> list[SessionGroup]:
        stmt = select(SessionGroup).order_by(SessionGroup.name)
        if owner_id is not None:
            stmt = stmt.where(SessionGroup.owner_id == owner_id)
        return list(db.scalars(stmt).all())

    def get_sessions_by_ids(self, db: Session, session_ids: list[int], owner_id: int | None = None) -> list[TelegramSession]:
        if not session_ids:
            return []
        stmt = select(TelegramSession).where(TelegramSession.id.in_(session_ids))
        if owner_id is not None:
            stmt = stmt.where(TelegramSession.owner_id == owner_id)
        return list(db.scalars(stmt).all())

    def export_session_files(
        self,
        db: Session,
        owner_id: int,
        session_ids: list[int] | None = None,
    ) -> tuple[Path, dict[str, int]]:
        if session_ids is not None:
            sessions = self.get_sessions_by_ids(db, list(dict.fromkeys(session_ids)), owner_id)
        else:
            sessions, _ = self.list_sessions(db, owner_id=owner_id, page_size=1_000_000)
        requested = len(set(session_ids)) if session_ids is not None else len(sessions)
        session_dir = Path(settings.session_dir).resolve()
        archive_fd, archive_name = tempfile.mkstemp(prefix="tg_sessions_", suffix=".zip")
        os.close(archive_fd)
        archive_path = Path(archive_name)
        exported = 0

        try:
            with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for session in sessions:
                    source = (session_dir / f"{session.session_name}.session").resolve()
                    if source.parent != session_dir or not source.is_file():
                        continue

                    snapshot_fd, snapshot_name = tempfile.mkstemp(suffix=".session")
                    os.close(snapshot_fd)
                    snapshot_path = Path(snapshot_name)
                    try:
                        with sqlite3.connect(f"{source.as_uri()}?mode=ro", uri=True, timeout=10) as source_db:
                            with sqlite3.connect(snapshot_path) as snapshot_db:
                                source_db.backup(snapshot_db)
                        archive.write(snapshot_path, arcname=f"{session.session_name}.session")
                        exported += 1
                    except sqlite3.Error:
                        continue
                    finally:
                        snapshot_path.unlink(missing_ok=True)
        except Exception:
            archive_path.unlink(missing_ok=True)
            raise

        if exported == 0:
            archive_path.unlink(missing_ok=True)
            raise ValueError("没有可导出的Session文件")

        return archive_path, {
            "requested": requested,
            "found": len(sessions),
            "exported": exported,
            "missing": requested - exported,
        }

    def create_group(self, db: Session, name: str, description: str | None = None, color: str = "blue", owner_id: int | None = None) -> SessionGroup:
        group = SessionGroup(owner_id=owner_id, name=name, description=description, color=color or "blue")
        db.add(group)
        db.commit()
        db.refresh(group)
        return group

    async def move_sessions(self, db: Session, session_ids: list[int], group_id: int | None, owner_id: int | None = None) -> int:
        target_group_id = None if group_id == 0 else group_id
        stmt = select(TelegramSession).where(TelegramSession.id.in_(session_ids))
        if owner_id is not None:
            stmt = stmt.where(TelegramSession.owner_id == owner_id)
        sessions = db.scalars(stmt).all()
        for session in sessions:
            session.group_id = target_group_id
            self.log(db, session.id, "move_group", f"Moved to group {target_group_id or 'none'}")
        db.commit()
        for session in sessions:
            db.refresh(session)
            await self.publish_status(session, "updated")
        return len(sessions)

    async def move_sessions_to_agent(self, db: Session, session_ids: list[int], kf_id: int | None, owner_id: int | None = None) -> int:
        target_kf_id = None if kf_id == 0 else kf_id
        stmt = select(TelegramSession).where(TelegramSession.id.in_(session_ids))
        if owner_id is not None:
            stmt = stmt.where(TelegramSession.owner_id == owner_id)
        sessions = db.scalars(stmt).all()
        for session in sessions:
            session.kf_id = target_kf_id
            self.log(db, session.id, "move_support_agent", f"Moved to support agent {target_kf_id or 'none'}")
        db.commit()
        for session in sessions:
            db.refresh(session)
            await self.publish_status(session, "updated")
        return len(sessions)

    async def import_sessions(self, db: Session, file: UploadFile, owner_id: int | None = None) -> dict[str, int]:
        content = await file.read()
        if (file.filename or "").lower().endswith(".session"):
            created = await self._import_session_file(db, file.filename or "", content, owner_id)
            skipped = 0 if created else 1
            await session_ws_manager.broadcast("*", {"event": "bulk_imported", "created": created, "skipped": skipped})
            return {"created": created, "skipped": skipped}

        rows = self._read_import_rows(file.filename or "", content)
        created = 0
        skipped = 0
        for row in rows:
            phone = self._pick_value(row, "phone", "手机号", "手机", "账号", "session", "session号", "session_name")
            username = self._pick_value(row, "username", "用户名", "昵称", "name")
            avatar = self._pick_value(row, "avatar", "头像", "avatar_url")
            group_id = self._pick_value(row, "group_id", "分组", "分组id")
            phone = self._normalize_phone(phone)
            if not phone:
                skipped += 1
                continue
            exists_stmt = select(TelegramSession).where(TelegramSession.phone == phone)
            if owner_id is not None:
                exists_stmt = exists_stmt.where(TelegramSession.owner_id == owner_id)
            exists = db.scalar(exists_stmt)
            if exists:
                skipped += 1
                continue
            self.create_session(
                db,
                {
                    "username": username or self._make_username(phone),
                    "phone": phone,
                    "avatar": avatar or None,
                    "group_id": self._normalize_group_id(group_id),
                },
                owner_id=owner_id,
            )
            created += 1
        await session_ws_manager.broadcast("*", {"event": "bulk_imported", "created": created, "skipped": skipped})
        return {"created": created, "skipped": skipped}

    async def health_check_once(self, db: Session, owner_id: int | None = None) -> dict[str, int]:
        stmt = select(TelegramSession).where(TelegramSession.status == SessionStatus.connected)
        if owner_id is not None:
            stmt = stmt.where(TelegramSession.owner_id == owner_id)
        sessions = db.scalars(stmt).all()
        checked = 0
        skipped = 0
        for session in sessions:
            if await redis_client.exists(f"marketing:session_lock:{session.id}"):
                skipped += 1
                continue
            checked += 1
            try:
                await session_command_bus.execute(session.id, "health", timeout=20)
                session.status = SessionStatus.connected
                session.health_status = "healthy"
                session.last_health_check_at = datetime.utcnow()
                session.error_message = None
                session.last_login_at = datetime.utcnow()
            except Exception as exc:
                session.status = SessionStatus.error
                session.health_status = "unhealthy"
                session.error_message = str(exc)
                session.last_health_check_at = datetime.utcnow()
            self.log(db, session.id, "health_check", session.health_status or "unknown")
            db.commit()
            db.refresh(session)
            await self.publish_status(session, "status_changed")
        return {"checked": checked, "skipped": skipped}

    async def check_bidirectional_status(self, db: Session, session_id: int, owner_id: int | None = None) -> TelegramSession:
        session = db.get(TelegramSession, session_id)
        if not session or (owner_id is not None and session.owner_id != owner_id):
            raise ValueError("Session not found")
        await self._require_runtime_online(db, session)
        if await redis_client.exists(f"marketing:session_lock:{session.id}"):
            raise ValueError("Session is busy sending a task")

        session.bidirectional_status = "checking"
        session.bidirectional_detail = None
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "bidirectional_checking")

        try:
            result = await session_command_bus.execute(session.id, "bidirectional_check", timeout=30)
            response_text = str(result.get("response") or "")
            session.bidirectional_detail = response_text[:10000]
            session.bidirectional_status = self._classify_bidirectional_response(response_text) if response_text else "timeout"
        except Exception as exc:
            session.bidirectional_status = "error"
            session.bidirectional_detail = str(exc)[:10000]

        session.last_bidirectional_check_at = datetime.utcnow()
        self.log(
            db,
            session.id,
            "bidirectional_check",
            f"{session.bidirectional_status}: {session.bidirectional_detail or '-'}"[:10000],
        )
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "bidirectional_checked")
        return session

    async def check_bidirectional_statuses(
        self,
        db: Session,
        session_ids: list[int],
        owner_id: int | None = None,
    ) -> dict[str, Any]:
        unique_ids = list(dict.fromkeys(session_ids))
        sessions = self.get_sessions_by_ids(db, unique_ids, owner_id)
        found_ids = {session.id for session in sessions}
        summary: dict[str, Any] = {
            "requested": len(unique_ids),
            "found": len(sessions),
            "accepted": 0,
            "accepted_ids": [],
            "skipped": len(unique_ids) - len(found_ids),
        }
        for session in sessions:
            lock_key = f"telegram:session:bidirectional-pending:{session.id}"
            lock_token = uuid.uuid4().hex
            acquired = await redis_client.set(
                lock_key,
                lock_token,
                nx=True,
                ex=self._BIDIRECTIONAL_QUEUE_LOCK_SECONDS,
            )
            if session.bidirectional_status == "checking" or not acquired:
                if acquired:
                    await self._release_queue_lock(lock_key, lock_token)
                summary["skipped"] += 1
                self.log(db, session.id, "bidirectional_check_skipped", "检测任务已在执行")
                db.commit()
                continue

            if await redis_client.exists(f"marketing:session_lock:{session.id}"):
                await self._release_queue_lock(lock_key, lock_token)
                summary["skipped"] += 1
                self.log(db, session.id, "bidirectional_check_skipped", "Session is busy sending a task")
                db.commit()
                continue

            session.bidirectional_status = "checking"
            session.bidirectional_detail = None
            self.log(db, session.id, "bidirectional_check_queued", "双向号检测任务已进入后台队列")
            db.commit()
            db.refresh(session)
            await self.publish_status(session, "bidirectional_checking")

            task = asyncio.create_task(
                self._run_queued_bidirectional_check(session.id, owner_id, lock_key, lock_token),
                name=f"bidirectional-check-{session.id}",
            )
            self._bidirectional_tasks.add(task)
            task.add_done_callback(self._bidirectional_tasks.discard)
            summary["accepted"] += 1
            summary["accepted_ids"].append(session.id)
        return summary

    async def _run_queued_bidirectional_check(
        self,
        session_id: int,
        owner_id: int | None,
        lock_key: str,
        lock_token: str,
    ) -> None:
        try:
            async with self._bidirectional_semaphore:
                db = SessionLocal()
                try:
                    session = self._owned_session(db, session_id, owner_id)
                    if not await redis_client.exists(f"telegram:session:runtime:{session_id}"):
                        if session.status == SessionStatus.connecting:
                            for _ in range(120):
                                if await redis_client.exists(f"telegram:session:runtime:{session_id}"):
                                    break
                                await asyncio.sleep(0.5)
                        if not await redis_client.exists(f"telegram:session:runtime:{session_id}"):
                            await self.connect_session(db, session)

                    result = await self.check_bidirectional_status(db, session_id, owner_id)
                    if result.bidirectional_status in {"restricted", "unknown"}:
                        await asyncio.sleep(1)
                        await self.check_bidirectional_status(db, session_id, owner_id)
                except Exception as exc:
                    db.rollback()
                    session = db.get(TelegramSession, session_id)
                    if session:
                        session.bidirectional_status = "error"
                        session.bidirectional_detail = str(exc)[:10000]
                        session.last_bidirectional_check_at = datetime.utcnow()
                        self.log(db, session.id, "bidirectional_check_failed", str(exc)[:10000])
                        db.commit()
                        db.refresh(session)
                        await self.publish_status(session, "bidirectional_checked")
                finally:
                    db.close()
        finally:
            await self._release_queue_lock(lock_key, lock_token)

    async def _release_queue_lock(self, key: str, token: str) -> None:
        await redis_client.eval(
            "if redis.call('get', KEYS[1]) == ARGV[1] then "
            "return redis.call('del', KEYS[1]) else return 0 end",
            1,
            key,
            token,
        )

    async def scan_contacts(self, db: Session, session_id: int, owner_id: int | None = None) -> TelegramSession:
        session = self._owned_session(db, session_id, owner_id)
        await self._require_runtime_online(db, session)
        result = await session_command_bus.execute(session.id, "contacts_scan", timeout=40)
        session.contact_count = int(result.get("contact_count", 0))
        session.contacts_scanned_at = datetime.utcnow()
        session.contact_scan_status = "success"
        session.contact_scan_detail = None
        self.log(db, session.id, "contacts_scan", f"识别到 {session.contact_count} 个通讯录好友")
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "contacts_updated")
        return session

    async def clear_contacts(self, db: Session, session_id: int, owner_id: int | None = None) -> TelegramSession:
        session = self._owned_session(db, session_id, owner_id)
        await self._require_runtime_online(db, session)
        result = await session_command_bus.execute(session.id, "contacts_clear", timeout=60)
        session.contact_count = 0
        session.contacts_scanned_at = datetime.utcnow()
        self.log(db, session.id, "contacts_clear", f"已清空通讯录，共删除 {int(result.get('deleted', 0))} 个好友")
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "contacts_updated")
        return session

    async def import_contacts(
        self,
        db: Session,
        session_id: int,
        phones: list[str],
        owner_id: int | None = None,
    ) -> TelegramSession:
        session = self._owned_session(db, session_id, owner_id)
        await self._require_runtime_online(db, session)
        result = await session_command_bus.execute(session.id, "contacts_import", {"phones": phones}, timeout=120)
        session.contact_count = int(result.get("contact_count", 0))
        session.contacts_scanned_at = datetime.utcnow()
        self.log(db, session.id, "contacts_import", f"导入 {len(phones)} 个号码，当前通讯录好友 {session.contact_count} 个")
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "contacts_updated")
        return session

    async def batch_contact_action(
        self,
        db: Session,
        session_ids: list[int],
        action: str,
        owner_id: int | None = None,
        phones: list[str] | None = None,
    ) -> dict[str, Any]:
        if action == "scan":
            return await self.queue_contact_scans(db, session_ids, owner_id)
        sessions = self.get_sessions_by_ids(db, session_ids, owner_id)
        connection_errors = await self._auto_connect_sessions([session.id for session in sessions], owner_id)
        found_ids = {session.id for session in sessions}
        missing_ids = [session_id for session_id in dict.fromkeys(session_ids) if session_id not in found_ids]
        summary: dict[str, Any] = {
            "requested": len(set(session_ids)),
            "found": len(sessions),
            "success": 0,
            "failed": len(missing_ids),
            "total_contacts": 0,
            "errors": [{"session_id": session_id, "session_name": None, "error": "Session not found"} for session_id in missing_ids],
        }
        for session in sessions:
            if session.id in connection_errors:
                error = connection_errors[session.id]
                summary["failed"] += 1
                summary["errors"].append({"session_id": session.id, "session_name": session.session_name, "error": error})
                self.log(db, session.id, f"contacts_{action}_failed", error)
                db.commit()
                continue
            try:
                if action == "scan":
                    result = await self.scan_contacts(db, session.id, owner_id)
                elif action == "clear":
                    result = await self.clear_contacts(db, session.id, owner_id)
                elif action == "import":
                    result = await self.import_contacts(db, session.id, phones or [], owner_id)
                else:
                    raise ValueError("Unsupported contact action")
                summary["success"] += 1
                summary["total_contacts"] += result.contact_count or 0
            except Exception as exc:
                db.rollback()
                summary["failed"] += 1
                summary["errors"].append({"session_id": session.id, "session_name": session.session_name, "error": str(exc)[:500]})
                self.log(db, session.id, f"contacts_{action}_failed", str(exc)[:500])
                db.commit()
        return summary

    async def queue_contact_scans(
        self,
        db: Session,
        session_ids: list[int],
        owner_id: int | None = None,
    ) -> dict[str, Any]:
        unique_ids = list(dict.fromkeys(session_ids))
        sessions = self.get_sessions_by_ids(db, unique_ids, owner_id)
        found_ids = {session.id for session in sessions}
        summary: dict[str, Any] = {
            "requested": len(unique_ids),
            "found": len(sessions),
            "accepted": 0,
            "accepted_ids": [],
            "skipped": len(unique_ids) - len(found_ids),
        }
        for session in sessions:
            lock_key = f"telegram:session:contacts-scan-pending:{session.id}"
            lock_token = uuid.uuid4().hex
            acquired = await redis_client.set(
                lock_key,
                lock_token,
                nx=True,
                ex=self._CONTACT_SCAN_QUEUE_LOCK_SECONDS,
            )
            if not acquired:
                summary["skipped"] += 1
                self.log(db, session.id, "contacts_scan_skipped", "通讯录识别任务已在执行")
                db.commit()
                continue
            if await redis_client.exists(f"marketing:session_lock:{session.id}"):
                await self._release_queue_lock(lock_key, lock_token)
                summary["skipped"] += 1
                self.log(db, session.id, "contacts_scan_skipped", "Session is busy sending a task")
                db.commit()
                continue

            session.contact_scan_status = "queued"
            session.contact_scan_detail = None
            self.log(db, session.id, "contacts_scan_queued", "通讯录识别任务已进入后台队列")
            db.commit()
            db.refresh(session)
            await self.publish_status(session, "contacts_scan_queued")

            task = asyncio.create_task(
                self._run_queued_contact_scan(session.id, owner_id, lock_key, lock_token),
                name=f"contacts-scan-{session.id}",
            )
            self._contact_scan_tasks.add(task)
            task.add_done_callback(self._contact_scan_tasks.discard)
            summary["accepted"] += 1
            summary["accepted_ids"].append(session.id)
        return summary

    async def _run_queued_contact_scan(
        self,
        session_id: int,
        owner_id: int | None,
        lock_key: str,
        lock_token: str,
    ) -> None:
        try:
            async with self._contact_scan_semaphore:
                db = SessionLocal()
                try:
                    session = self._owned_session(db, session_id, owner_id)
                    session.contact_scan_status = "scanning"
                    db.commit()
                    db.refresh(session)
                    await self.publish_status(session, "contacts_scanning")

                    if not await redis_client.exists(f"telegram:session:runtime:{session_id}"):
                        await self.connect_session(db, session)
                    await self.scan_contacts(db, session_id, owner_id)
                except Exception as exc:
                    db.rollback()
                    session = db.get(TelegramSession, session_id)
                    if session:
                        session.contact_scan_status = "failed"
                        session.contact_scan_detail = str(exc)[:1000]
                        self.log(db, session.id, "contacts_scan_failed", str(exc)[:1000])
                        db.commit()
                        db.refresh(session)
                        await self.publish_status(session, "contacts_updated")
                finally:
                    db.close()
        finally:
            await self._release_queue_lock(lock_key, lock_token)

    async def distribute_import_contacts(
        self,
        db: Session,
        session_ids: list[int],
        phones: list[str],
        per_session_limit: int,
        owner_id: int | None = None,
    ) -> dict[str, Any]:
        if per_session_limit < 1 or per_session_limit > 10000:
            raise ValueError("每个Session导入数量必须在1到10000之间")

        unique_ids = list(dict.fromkeys(session_ids))
        sessions_by_id = {
            session.id: session for session in self.get_sessions_by_ids(db, unique_ids, owner_id)
        }
        connection_errors = await self._auto_connect_sessions(list(sessions_by_id), owner_id)
        summary: dict[str, Any] = {
            "requested": len(unique_ids),
            "found": len(sessions_by_id),
            "success": 0,
            "failed": 0,
            "allocated_count": 0,
            "remaining_count": 0,
            "remaining_phones": [],
            "total_contacts": 0,
            "errors": [],
        }
        cursor = 0
        for session_id in unique_ids:
            session = sessions_by_id.get(session_id)
            if not session:
                summary["failed"] += 1
                summary["errors"].append({"session_id": session_id, "session_name": None, "error": "Session not found"})
                continue
            if session_id in connection_errors:
                error = connection_errors[session_id]
                summary["failed"] += 1
                summary["errors"].append({"session_id": session_id, "session_name": session.session_name, "error": error})
                self.log(db, session.id, "contacts_import_failed", error)
                db.commit()
                continue
            assigned = phones[cursor : cursor + per_session_limit]
            if not assigned:
                break
            cursor += len(assigned)
            summary["allocated_count"] += len(assigned)
            try:
                result = await self.import_contacts(db, session.id, assigned, owner_id)
                summary["success"] += 1
                summary["total_contacts"] += result.contact_count or 0
            except Exception as exc:
                db.rollback()
                summary["failed"] += 1
                summary["errors"].append({
                    "session_id": session.id,
                    "session_name": session.session_name,
                    "assigned_count": len(assigned),
                    "error": str(exc)[:500],
                })

        summary["remaining_phones"] = phones[cursor:]
        summary["remaining_count"] = len(summary["remaining_phones"])
        return summary

    def parse_contact_phones(self, content: bytes) -> list[str]:
        if len(content) > 5 * 1024 * 1024:
            raise ValueError("TXT文件不能超过5MB")
        phones = parse_targets(content, "phone")
        if not phones:
            raise ValueError("TXT文件中没有有效手机号")
        if len(phones) > 10000:
            raise ValueError("单次最多导入10000个手机号")
        return phones

    def _owned_session(self, db: Session, session_id: int, owner_id: int | None) -> TelegramSession:
        session = db.get(TelegramSession, session_id)
        if not session or (owner_id is not None and session.owner_id != owner_id):
            raise ValueError("Session not found")
        return session

    async def _require_runtime_online(self, db: Session, session: TelegramSession) -> None:
        if await redis_client.exists(f"telegram:session:runtime:{session.id}"):
            return
        result = await self.connect_session(db, session)
        if result.status != SessionStatus.connected or not await redis_client.exists(f"telegram:session:runtime:{session.id}"):
            raise ValueError(
                f"Session {session.session_name} 自动连接失败：{result.error_message or '未建立真实在线心跳'}"
            )

    async def _auto_connect_sessions(
        self, session_ids: list[int], owner_id: int | None,
    ) -> dict[int, str]:
        """Connect offline Sessions concurrently and return per-Session errors."""
        connect_limit = asyncio.Semaphore(10)

        async def connect_one(session_id: int) -> tuple[int, str | None]:
            if await redis_client.exists(f"telegram:session:runtime:{session_id}"):
                return session_id, None
            local_db = SessionLocal()
            try:
                session = self._owned_session(local_db, session_id, owner_id)
                async with connect_limit:
                    result = await self.connect_session(local_db, session)
                if result.status == SessionStatus.connected and await redis_client.exists(
                    f"telegram:session:runtime:{session_id}"
                ):
                    return session_id, None
                return session_id, result.error_message or "未建立真实在线心跳"
            except Exception as exc:
                return session_id, str(exc)[:500]
            finally:
                local_db.close()

        results = await asyncio.gather(*[connect_one(session_id) for session_id in dict.fromkeys(session_ids)])
        return {session_id: error for session_id, error in results if error}

    def _telegram_contact_count(self, result: Any) -> int:
        contacts = getattr(result, "contacts", None)
        if contacts is not None:
            return len(contacts)
        return len(getattr(result, "users", []) or [])

    def list_logs(self, db: Session, session_id: int | None = None, limit: int = 100, owner_id: int | None = None) -> list[SessionLog]:
        stmt = select(SessionLog).order_by(SessionLog.created_at.desc()).limit(limit)
        if owner_id is not None:
            stmt = stmt.join(TelegramSession, SessionLog.session_id == TelegramSession.id).where(TelegramSession.owner_id == owner_id)
        if session_id:
            stmt = stmt.where(SessionLog.session_id == session_id)
        return list(db.scalars(stmt).all())

    def count_sent_messages(self, db: Session, session_id: int) -> int:
        return db.scalar(
            select(func.count(SessionTaskLog.id)).where(SessionTaskLog.session_id == session_id, SessionTaskLog.status == "success")
        ) or 0

    def list_task_logs(self, db: Session, session_id: int, limit: int = 100) -> list[SessionTaskLog]:
        stmt = (
            select(SessionTaskLog)
            .where(SessionTaskLog.session_id == session_id)
            .order_by(SessionTaskLog.created_at.desc())
            .limit(limit)
        )
        return list(db.scalars(stmt).all())

    def log(self, db: Session, session_id: int | None, action: str, message: str) -> None:
        db.add(SessionLog(session_id=session_id, action=action, message=message))

    async def publish_status(self, session: TelegramSession, event: str) -> None:
        payload = {"event": event, "session": self.serialize_session(session)}
        await session_ws_manager.broadcast(str(session.id), payload)
        await session_ws_manager.broadcast("*", payload)

    def serialize_session(
        self,
        session: TelegramSession,
        runtime: dict[str, Any] | None = None,
        include_runtime: bool = False,
        proxy: Any = None,
        sent_count: int | None = None,
        resolve_related: bool = True,
    ) -> dict[str, Any]:
        db = object_session(session)
        if resolve_related:
            proxy = proxy_service.get_proxy_for_session(db, session, require_active=False) if db else None
        payload = {
            "id": session.id,
            "username": session.username,
            "avatar": session.avatar,
            "phone": session.phone,
            "session_name": session.session_name,
            "status": session.status.value if hasattr(session.status, "value") else session.status,
            "last_login_at": session.last_login_at.isoformat() if session.last_login_at else None,
            "last_health_check_at": session.last_health_check_at.isoformat() if session.last_health_check_at else None,
            "health_status": session.health_status,
            "error_message": session.error_message,
            "bidirectional_status": session.bidirectional_status or "unchecked",
            "bidirectional_detail": session.bidirectional_detail,
            "last_bidirectional_check_at": session.last_bidirectional_check_at.isoformat() if session.last_bidirectional_check_at else None,
            "contact_count": session.contact_count,
            "contacts_scanned_at": session.contacts_scanned_at.isoformat() if session.contacts_scanned_at else None,
            "contact_scan_status": session.contact_scan_status or "idle",
            "contact_scan_detail": session.contact_scan_detail,
            "group_id": session.group_id,
            "group_name": session.group.name if session.group else None,
            "group_color": session.group.color if session.group else None,
            "kf_id": session.kf_id,
            "kf_name": session.support_agent.name if session.support_agent else None,
            "kf_color": session.support_agent.color if session.support_agent else None,
            "proxy_id": proxy.id if proxy else None,
            "proxy_name": proxy.name if proxy else None,
            "proxy_color": proxy.color if proxy else None,
            "proxy_status": proxy.status if proxy else None,
            "sent_count": self.count_sent_messages(db, session.id) if db and sent_count is None else (sent_count or 0),
            "created_at": session.created_at.isoformat() if session.created_at else None,
            "updated_at": session.updated_at.isoformat() if session.updated_at else None,
        }
        if include_runtime:
            payload.update({
                "runtime_status": (runtime or {}).get("status", "offline"),
                "runtime_worker": (runtime or {}).get("worker"),
                "runtime_last_heartbeat": (runtime or {}).get("last_heartbeat"),
            })
        return payload

    async def _sync_telegram_profile(self, client: Any, session: TelegramSession, me: Any) -> None:
        display_name = " ".join(part for part in [getattr(me, "first_name", None), getattr(me, "last_name", None)] if part)
        session.username = me.username or display_name or session.username
        session.phone = me.phone or session.phone
        avatar_path = await self._download_avatar(client, session.session_name, me)
        if avatar_path:
            session.avatar = avatar_path

    async def _download_avatar(self, client: Any, session_name: str, me: Any) -> str | None:
        avatar_dir = Path("static") / "avatars"
        avatar_dir.mkdir(parents=True, exist_ok=True)
        avatar_file = avatar_dir / f"{session_name}.jpg"
        try:
            downloaded = await client.download_profile_photo(me, file=str(avatar_file))
        except Exception:
            return None
        if not downloaded:
            return None
        return f"/static/avatars/{avatar_file.name}"

    def _make_session_name(self, phone: str) -> str:
        return "tg_" + "".join(ch for ch in phone if ch.isalnum())

    async def _import_session_file(self, db: Session, filename: str, content: bytes, owner_id: int | None = None) -> int:
        session_name = self._session_name_from_filename(filename)
        if not session_name:
            return 0
        exists_stmt = select(TelegramSession).where(TelegramSession.session_name == session_name)
        if owner_id is not None:
            exists_stmt = exists_stmt.where(TelegramSession.owner_id == owner_id)
        exists = db.scalar(exists_stmt)
        if exists:
            return 0

        session_dir = Path(settings.session_dir)
        session_dir.mkdir(parents=True, exist_ok=True)
        session_path = session_dir / f"{session_name}.session"
        session_path.write_bytes(content)

        username = session_name[:100]
        phone = self._normalize_phone(session_name) or session_name[:32]

        phone_exists_stmt = select(TelegramSession).where(TelegramSession.phone == phone)
        if owner_id is not None:
            phone_exists_stmt = phone_exists_stmt.where(TelegramSession.owner_id == owner_id)
        phone_exists = db.scalar(phone_exists_stmt)
        if phone_exists:
            session_path.unlink(missing_ok=True)
            return 0

        session = self.create_session(db, {"username": username, "phone": phone, "session_name": session_name}, owner_id=owner_id)
        session.status = SessionStatus.disconnected
        session.health_status = "unchecked"
        session.error_message = None
        session.last_health_check_at = datetime.utcnow()
        self.log(db, session.id, "import_session_file", f"Imported {filename}")
        db.commit()
        db.refresh(session)
        await self.publish_status(session, "updated")
        return 1

    def _session_name_from_filename(self, filename: str) -> str:
        stem = Path(filename).stem.strip()
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", stem)[:150]

    def _read_import_rows(self, filename: str, content: bytes) -> list[dict[str, Any]]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".txt":
            return self._read_plain_text_rows(content)

        if filename.lower().endswith(".csv"):
            text = content.decode("utf-8-sig")
            return self._read_csv_rows(text)

        workbook = load_workbook(io.BytesIO(content), read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        return self._read_tabular_rows(rows)

    def _read_csv_rows(self, text: str) -> list[dict[str, Any]]:
        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect))
        return self._read_tabular_rows(rows)

    def _read_plain_text_rows(self, content: bytes) -> list[dict[str, Any]]:
        text = content.decode("utf-8-sig", errors="ignore")
        rows: list[dict[str, Any]] = []
        for line in text.splitlines():
            value = line.strip()
            if not value:
                continue
            parts = re.split(r"[\s,，;；]+", value)
            if parts:
                rows.append({"phone": parts[0], "username": parts[1] if len(parts) > 1 else ""})
        return rows

    def _read_tabular_rows(self, rows: list[Any]) -> list[dict[str, Any]]:
        normalized_rows = [list(row) for row in rows if any(cell not in (None, "") for cell in row)]
        if not normalized_rows:
            return []

        first_row = [str(cell).strip() if cell is not None else "" for cell in normalized_rows[0]]
        if self._looks_like_header(first_row):
            headers = first_row
            data_rows = normalized_rows[1:]
            return [dict(zip(headers, row)) for row in data_rows]

        return [
            {
                "phone": row[0] if len(row) > 0 else "",
                "username": row[1] if len(row) > 1 else "",
                "avatar": row[2] if len(row) > 2 else "",
                "group_id": row[3] if len(row) > 3 else "",
            }
            for row in normalized_rows
        ]

    def _looks_like_header(self, row: list[str]) -> bool:
        known_headers = {
            "phone",
            "手机号",
            "手机",
            "账号",
            "session",
            "session号",
            "session_name",
            "username",
            "用户名",
            "昵称",
            "name",
            "avatar",
            "头像",
            "avatar_url",
            "group_id",
            "分组",
            "分组id",
        }
        return any(cell.lower() in known_headers for cell in row)

    def _pick_value(self, row: dict[str, Any], *keys: str) -> str:
        normalized = {str(key).strip().lower(): value for key, value in row.items()}
        for key in keys:
            value = normalized.get(key.lower())
            if value not in (None, ""):
                return str(value).strip()
        return ""

    def _normalize_phone(self, value: str) -> str:
        value = value.strip()
        if not value:
            return ""
        value = re.sub(r"\.0$", "", value)
        match = re.search(r"\+?\d[\d\s().-]{4,}\d", value)
        if not match:
            return ""
        phone = re.sub(r"[\s().-]+", "", match.group(0))
        return phone[:32]

    def _make_username(self, phone: str) -> str:
        suffix = "".join(ch for ch in phone if ch.isdigit())[-6:] or "session"
        return f"session_{suffix}"

    def _normalize_group_id(self, value: str) -> int | None:
        if not value:
            return None
        try:
            return int(float(value))
        except ValueError:
            return None

    async def _await_spam_bot_reply(self, client: Any, sent_message_id: int, sent_message_date: datetime | None) -> str:
        deadline = asyncio.get_running_loop().time() + 10
        normalized_sent_at = self._normalize_message_datetime(sent_message_date)
        while True:
            async for message in client.iter_messages("@SpamBot", limit=5):
                if getattr(message, "out", False):
                    continue
                message_id = getattr(message, "id", 0) or 0
                message_date = self._normalize_message_datetime(getattr(message, "date", None))
                if message_id <= sent_message_id:
                    continue
                if normalized_sent_at and message_date and message_date < normalized_sent_at:
                    continue
                return message.text or message.raw_text or ""
            if asyncio.get_running_loop().time() >= deadline:
                raise asyncio.TimeoutError()
            await asyncio.sleep(0.5)

    def _normalize_message_datetime(self, value: datetime | None) -> datetime | None:
        if not value:
            return None
        if value.tzinfo:
            return value.astimezone().replace(tzinfo=None)
        return value

    def _classify_bidirectional_response(self, response_text: str) -> str:
        normalized = (response_text or "").strip().lower()
        if any(marker in normalized for marker in self._SPAM_BOT_NORMAL_MARKERS):
            return "normal"
        if any(marker in normalized for marker in self._SPAM_BOT_BLOCKED_MARKERS):
            return "blocked"
        if any(marker in normalized for marker in self._SPAM_BOT_RESTRICTED_MARKERS):
            return "restricted"
        return "unknown"


session_service = SessionService()


async def health_check_loop() -> None:
    while True:
        await asyncio.sleep(settings.health_check_interval_seconds)
        from app.core.database import SessionLocal

        db = SessionLocal()
        try:
            await session_service.health_check_once(db)
        finally:
            db.close()
